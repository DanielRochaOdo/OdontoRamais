from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = Path(r"C:\Users\daniel.rocha\Downloads\ramais.xlsx")
DEFAULT_CSV_TARGET = ROOT / "supabase" / "seed_ramais.csv"
DEFAULT_SQL_TARGET = ROOT / "supabase" / "migrations" / "20260727124500_seed_ramais_from_xlsx.sql"


def normalize(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text


def normalize_setor(value: object) -> str:
    return normalize(value).lstrip("*").strip()


def to_records(source: Path) -> list[dict[str, str]]:
    workbook = load_workbook(source)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    header = rows[0]
    records: list[dict[str, str]] = []

    for row in rows[1:]:
        data = dict(zip(header, row))
        records.append(
                {
                    "nome": normalize(data.get("Nome")),
                    "numero": normalize(data.get("Numero")),
                    "cargo": normalize(data.get("Cargo")),
                    "setor": normalize_setor(data.get("Setor")),
                    "email": "",
                    "observacoes": "",
                    "ativo": "true",
                }
        )

    return records


def write_csv(records: list[dict[str, str]], target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=["nome", "numero", "cargo", "setor", "email", "observacoes", "ativo"],
        )
        writer.writeheader()
        writer.writerows(records)


def sql_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def write_sql(records: list[dict[str, str]], target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    values = ",\n".join(
        (
            f"  ({sql_literal(row['nome'])}, {sql_literal(row['numero'])}, {sql_literal(row['cargo'])}, "
            f"{sql_literal(row['setor'])}, null, null, true)"
        )
        for row in records
    )
    sql = (
        "truncate table public.ramais;\n\n"
        "insert into public.ramais (nome, numero, cargo, setor, email, observacoes, ativo)\n"
        "values\n"
        f"{values};\n"
    )
    target.write_text(sql, encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--csv-target", type=Path, default=DEFAULT_CSV_TARGET)
    parser.add_argument("--sql-target", type=Path, default=DEFAULT_SQL_TARGET)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    records = to_records(args.source)
    write_csv(records, args.csv_target)
    write_sql(records, args.sql_target)
    print(f"CSV: {args.csv_target}")
    print(f"SQL: {args.sql_target}")
    print(f"Registros: {len(records)}")


if __name__ == "__main__":
    main()
